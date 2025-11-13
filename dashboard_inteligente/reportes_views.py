import json
import logging
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from dashboard_inteligente.views import GenerarPrediccionesView
from productos.models import Producto
from ventas_carrito.models import Venta

logger = logging.getLogger(__name__)


@method_decorator(csrf_exempt, name='dispatch')
class ExportarDashboardVentasView(View):
    """Exportar reporte del Dashboard de Ventas en PDF o Excel"""
    
    def get(self, request):
        """Exportar dashboard de ventas"""
        try:
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            formato = request.GET.get('formato', 'pdf')  # pdf o excel
            periodo = request.GET.get('periodo', '12')  # meses a incluir
            
            # Obtener datos del dashboard directamente
            from ventas_carrito.historial_views import DashboardStatsView
            stats_view = DashboardStatsView()
            stats_response = stats_view.get(request)
            
            if stats_response.status_code != 200:
                return JsonResponse({
                    'success': False,
                    'message': 'Error al obtener datos del dashboard'
                }, status=500)
            
            stats_data = json.loads(stats_response.content)
            
            if formato == 'pdf':
                return self._generar_pdf(stats_data, periodo)
            elif formato == 'excel':
                return self._generar_excel(stats_data, periodo)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato no soportado. Use pdf o excel'
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error en ExportarDashboardVentasView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar reporte: {str(e)}'
            }, status=500)
    
    def _generar_pdf(self, stats_data, periodo):
        """Generar PDF del dashboard de ventas"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066FF'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        # Título
        story.append(Paragraph("Reporte de Dashboard de Ventas", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Información del reporte
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        info_data = [
            ['Fecha de Generación:', fecha_actual],
            ['Período Analizado:', f'Últimos {periodo} meses'],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Estadísticas principales
        if stats_data.get('success') and stats_data.get('stats'):
            stats = stats_data['stats']
            
            story.append(Paragraph("Estadísticas Principales", heading_style))
            
            stats_data_table = [
                ['Métrica', 'Valor', 'Cambio %', 'Tendencia'],
            ]
            
            if stats.get('ventas_mes'):
                stats_data_table.append([
                    'Ventas del Mes',
                    f"Bs. {stats['ventas_mes']['value']:,.2f}",
                    f"{stats['ventas_mes']['change']:+.1f}%",
                    stats['ventas_mes']['trend'].upper()
                ])
            
            if stats.get('total_pedidos'):
                stats_data_table.append([
                    'Total Pedidos',
                    str(stats['total_pedidos']['value']),
                    f"{stats['total_pedidos']['change']:+.1f}%",
                    stats['total_pedidos']['trend'].upper()
                ])
            
            if stats.get('nuevos_clientes'):
                stats_data_table.append([
                    'Nuevos Clientes',
                    str(stats['nuevos_clientes']['value']),
                    f"{stats['nuevos_clientes']['change']:+.1f}%",
                    stats['nuevos_clientes']['trend'].upper()
                ])
            
            if stats.get('productos_activos'):
                stats_data_table.append([
                    'Productos Activos',
                    str(stats['productos_activos']['value']),
                    f"{stats['productos_activos']['change']:+.1f}%",
                    stats['productos_activos']['trend'].upper()
                ])
            
            stats_table = Table(stats_data_table, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Ventas mensuales
            if stats.get('ventas_mensuales'):
                story.append(Paragraph("Ventas Mensuales", heading_style))
                
                ventas_mensuales = stats['ventas_mensuales']
                labels = ventas_mensuales.get('labels', [])
                values = ventas_mensuales.get('values', [])
                
                if labels and values:
                    ventas_table_data = [['Mes', 'Ventas (Bs.)']]
                    for i, label in enumerate(labels):
                        if i < len(values):
                            ventas_table_data.append([
                                label,
                                f"Bs. {values[i]:,.2f}"
                            ])
                    
                    ventas_table = Table(ventas_table_data, colWidths=[3*inch, 3*inch])
                    ventas_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                    ]))
                    story.append(ventas_table)
                    story.append(Spacer(1, 0.3*inch))
            
            # Productos top
            if stats.get('productos_top'):
                story.append(Paragraph("Productos Más Vendidos", heading_style))
                
                productos_top = stats['productos_top'][:10]  # Top 10
                productos_table_data = [['Producto', 'Unidades Vendidas', 'Total (Bs.)']]
                
                for prod in productos_top:
                    productos_table_data.append([
                        prod.get('nombre', 'N/A'),
                        str(prod.get('cantidad', 0)),
                        f"Bs. {prod.get('total', 0):,.2f}"
                    ])
                
                productos_table = Table(productos_table_data, colWidths=[3*inch, 2*inch, 2*inch])
                productos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ]))
                story.append(productos_table)
        
        # Pie de página
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(
            f"<i>Reporte generado el {fecha_actual} - SmartSales365</i>",
            styles['Normal']
        ))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="dashboard_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _generar_excel(self, stats_data, periodo):
        """Generar Excel del dashboard de ventas"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Ventas"
        
        # Título
        ws['A1'] = "Reporte de Dashboard de Ventas"
        ws['A1'].font = Font(bold=True, size=18, color="0066FF")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Información del reporte
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        ws['A3'] = 'Fecha de Generación:'
        ws['B3'] = fecha_actual
        ws['A4'] = 'Período Analizado:'
        ws['B4'] = f'Últimos {periodo} meses'
        
        row = 6
        
        if stats_data.get('success') and stats_data.get('stats'):
            stats = stats_data['stats']
            
            # Estadísticas principales
            ws[f'A{row}'] = 'Estadísticas Principales'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            headers = ['Métrica', 'Valor', 'Cambio %', 'Tendencia']
            ws.append(headers)
            
            # Estilo para encabezados
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Datos de estadísticas
            if stats.get('ventas_mes'):
                ws.append([
                    'Ventas del Mes',
                    f"Bs. {stats['ventas_mes']['value']:,.2f}",
                    f"{stats['ventas_mes']['change']:+.1f}%",
                    stats['ventas_mes']['trend'].upper()
                ])
                row += 1
            
            if stats.get('total_pedidos'):
                ws.append([
                    'Total Pedidos',
                    stats['total_pedidos']['value'],
                    f"{stats['total_pedidos']['change']:+.1f}%",
                    stats['total_pedidos']['trend'].upper()
                ])
                row += 1
            
            if stats.get('nuevos_clientes'):
                ws.append([
                    'Nuevos Clientes',
                    stats['nuevos_clientes']['value'],
                    f"{stats['nuevos_clientes']['change']:+.1f}%",
                    stats['nuevos_clientes']['trend'].upper()
                ])
                row += 1
            
            if stats.get('productos_activos'):
                ws.append([
                    'Productos Activos',
                    stats['productos_activos']['value'],
                    f"{stats['productos_activos']['change']:+.1f}%",
                    stats['productos_activos']['trend'].upper()
                ])
                row += 1
            
            row += 2
            
            # Ventas mensuales
            if stats.get('ventas_mensuales'):
                ws[f'A{row}'] = 'Ventas Mensuales'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ventas_mensuales = stats['ventas_mensuales']
                labels = ventas_mensuales.get('labels', [])
                values = ventas_mensuales.get('values', [])
                
                ws.append(['Mes', 'Ventas (Bs.)'])
                # Estilo encabezado
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for i, label in enumerate(labels):
                    if i < len(values):
                        ws.append([label, values[i]])
                        row += 1
                
                # Crear gráfico de barras
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Ventas Mensuales"
                chart.y_axis.title = 'Ventas (Bs.)'
                chart.x_axis.title = 'Mes'
                
                data = Reference(ws, min_col=2, min_row=row-len(labels), max_row=row-1)
                cats = Reference(ws, min_col=1, min_row=row-len(labels), max_row=row-1)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"E{row-len(labels)}")
                row += len(labels) + 5
            
            # Productos top
            if stats.get('productos_top'):
                ws[f'A{row}'] = 'Productos Más Vendidos'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                productos_top = stats['productos_top'][:10]
                ws.append(['Producto', 'Unidades Vendidas', 'Total (Bs.)'])
                
                # Estilo encabezado
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for prod in productos_top:
                    ws.append([
                        prod.get('nombre', 'N/A'),
                        prod.get('cantidad', 0),
                        prod.get('total', 0)
                    ])
                    row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dashboard_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response


@method_decorator(csrf_exempt, name='dispatch')
class ExportarPrediccionesView(View):
    """Exportar reporte de Predicciones de IA en PDF o Excel"""
    
    def get(self, request):
        """Exportar predicciones"""
        try:
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            formato = request.GET.get('formato', 'pdf')  # pdf o excel
            
            # Obtener predicciones - usar listar predicciones
            from reportes_dinamicos.models import PrediccionVenta
            from productos.models import Categoria
            
            # Obtener todas las predicciones recientes
            predicciones_queryset = PrediccionVenta.objects.all().order_by('-fecha_prediccion')[:100]
            
            predicciones_list = []
            for pred in predicciones_queryset:
                categoria_data = None
                if pred.categoria:
                    categoria_data = {
                        'id_categoria': pred.categoria.id_categoria,
                        'nombre': pred.categoria.nombre
                    }
                
                predicciones_list.append({
                    'fecha_prediccion': pred.fecha_prediccion.strftime('%Y-%m-%d') if pred.fecha_prediccion else None,
                    'valor_predicho': float(pred.valor_predicho) if pred.valor_predicho else 0,
                    'confianza': float(pred.confianza) if pred.confianza else 0,
                    'categoria': categoria_data
                })
            
            # Calcular resumen
            total_predicciones = len(predicciones_list)
            total_valor_predicho = sum(p['valor_predicho'] for p in predicciones_list)
            confianza_promedio = sum(p['confianza'] for p in predicciones_list) / total_predicciones if total_predicciones > 0 else 0
            
            predicciones_data = {
                'success': True,
                'predicciones': predicciones_list,
                'resumen': {
                    'total_predicciones': total_predicciones,
                    'total_valor_predicho': total_valor_predicho,
                    'confianza_promedio': confianza_promedio,
                    'tendencias': {
                        'factor_crecimiento': 0,  # Se calcularía si hay datos históricos
                        'promedio_mensual_historico': 0
                    }
                }
            }
            
            # Obtener estado del modelo
            from dashboard_inteligente.views import EstadoModeloView
            modelo_view = EstadoModeloView()
            modelo_response = modelo_view.get(request)
            modelo_data = json.loads(modelo_response.content) if modelo_response.status_code == 200 else None
            
            if formato == 'pdf':
                return self._generar_pdf(predicciones_data, modelo_data)
            elif formato == 'excel':
                return self._generar_excel(predicciones_data, modelo_data)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato no soportado. Use pdf o excel'
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error en ExportarPrediccionesView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar reporte: {str(e)}'
            }, status=500)
    
    def _generar_pdf(self, predicciones_data, modelo_data):
        """Generar PDF de predicciones"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#8B5CF6'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        # Título
        story.append(Paragraph("Reporte de Predicciones de IA", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        info_data = [
            ['Fecha de Generación:', fecha_actual],
        ]
        
        if modelo_data and modelo_data.get('modelo'):
            modelo = modelo_data['modelo']
            info_data.append(['Modelo de IA:', f"{modelo.get('nombre', 'N/A')} v{modelo.get('version', 'N/A')}"])
            if modelo.get('r2_score'):
                info_data.append(['R² Score:', f"{modelo['r2_score']:.3f}"])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Resumen de predicciones
        if predicciones_data.get('success') and predicciones_data.get('resumen'):
            resumen = predicciones_data['resumen']
            story.append(Paragraph("Resumen de Predicciones", heading_style))
            
            resumen_data = [
                ['Total Predicciones:', str(resumen.get('total_predicciones', 0))],
                ['Total Valor Predicho:', f"Bs. {resumen.get('total_valor_predicho', 0):,.2f}"],
                ['Confianza Promedio:', f"{resumen.get('confianza_promedio', 0) * 100:.1f}%"],
            ]
            
            if resumen.get('tendencias'):
                tendencias = resumen['tendencias']
                resumen_data.append(['Factor de Crecimiento:', f"{tendencias.get('factor_crecimiento', 0):+.1f}%"])
                resumen_data.append(['Promedio Mensual Histórico:', f"Bs. {tendencias.get('promedio_mensual_historico', 0):,.2f}"])
            
            resumen_table = Table(resumen_data, colWidths=[3*inch, 3*inch])
            resumen_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(resumen_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Lista de predicciones
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            story.append(Paragraph("Predicciones Generadas", heading_style))
            
            predicciones_table_data = [['Fecha', 'Valor Predicho (Bs.)', 'Confianza', 'Categoría']]
            
            for pred in predicciones[:50]:  # Limitar a 50 para PDF
                predicciones_table_data.append([
                    pred.get('fecha_prediccion', 'N/A')[:10],
                    f"{pred.get('valor_predicho', 0):,.2f}",
                    f"{pred.get('confianza', 0) * 100:.1f}%",
                    pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                ])
            
            predicciones_table = Table(predicciones_table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
            predicciones_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(predicciones_table)
            
            if len(predicciones) > 50:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(
                    f"<i>Nota: Se muestran 50 de {len(predicciones)} predicciones totales.</i>",
                    styles['Normal']
                ))
        
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(
            f"<i>Reporte generado el {fecha_actual} - SmartSales365</i>",
            styles['Normal']
        ))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="predicciones_ia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _generar_excel(self, predicciones_data, modelo_data):
        """Generar Excel de predicciones"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Predicciones IA"
        
        # Título
        ws['A1'] = "Reporte de Predicciones de IA"
        ws['A1'].font = Font(bold=True, size=18, color="8B5CF6")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        ws['A3'] = 'Fecha de Generación:'
        ws['B3'] = fecha_actual
        
        row = 5
        
        if modelo_data and modelo_data.get('modelo'):
            modelo = modelo_data['modelo']
            ws[f'A{row}'] = 'Información del Modelo'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Campo', 'Valor'])
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws.append(['Nombre', modelo.get('nombre', 'N/A')])
            row += 1
            ws.append(['Versión', modelo.get('version', 'N/A')])
            row += 1
            if modelo.get('r2_score'):
                ws.append(['R² Score', f"{modelo['r2_score']:.3f}"])
                row += 1
            row += 2
        
        # Resumen
        if predicciones_data.get('success') and predicciones_data.get('resumen'):
            resumen = predicciones_data['resumen']
            ws[f'A{row}'] = 'Resumen de Predicciones'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Métrica', 'Valor'])
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws.append(['Total Predicciones', resumen.get('total_predicciones', 0)])
            row += 1
            ws.append(['Total Valor Predicho', f"Bs. {resumen.get('total_valor_predicho', 0):,.2f}"])
            row += 1
            ws.append(['Confianza Promedio', f"{resumen.get('confianza_promedio', 0) * 100:.1f}%"])
            row += 1
            
            if resumen.get('tendencias'):
                tendencias = resumen['tendencias']
                ws.append(['Factor de Crecimiento', f"{tendencias.get('factor_crecimiento', 0):+.1f}%"])
                row += 1
                ws.append(['Promedio Mensual Histórico', f"Bs. {tendencias.get('promedio_mensual_historico', 0):,.2f}"])
                row += 1
            
            row += 2
        
        # Predicciones
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            ws[f'A{row}'] = 'Predicciones Generadas'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Fecha', 'Valor Predicho (Bs.)', 'Confianza (%)', 'Categoría'])
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            for pred in predicciones:
                ws.append([
                    pred.get('fecha_prediccion', 'N/A')[:10],
                    pred.get('valor_predicho', 0),
                    pred.get('confianza', 0) * 100,
                    pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                ])
                row += 1
            
            # Crear gráfico de línea
            if len(predicciones) > 0:
                chart = LineChart()
                chart.title = "Evolución de Predicciones"
                chart.y_axis.title = 'Valor Predicho (Bs.)'
                chart.x_axis.title = 'Fecha'
                
                data = Reference(ws, min_col=2, min_row=row-len(predicciones), max_row=row-1)
                cats = Reference(ws, min_col=1, min_row=row-len(predicciones), max_row=row-1)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"F{row-len(predicciones)}")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="predicciones_ia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

